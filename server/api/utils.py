# this file contains functions that are useful and used in many different places
from oauthlib.oauth2 import WebApplicationClient
import os
import re
import requests
from server.api.extensions import fake, db
from server.api.constants import LITTLE_QUESTION_SET_ID, BIG_QUESTION_SET_ID, NONE, LOW, MEDIUM, HIGH
from server.api.models import Big_Question_Set, Little_Question_Set, Big_Answer, Little_Answer
from server.api.gmail_sender import send_little_confirm_emails, send_big_confirm_emails
from server.api.gmail_sender import send_event_emails
import openpyxl

#OAuth 2 client setup
# for google authentication
GOOGLE_CLIENT_ID = os.getenv("GOOGLE_CLIENT_ID", None)
GOOGLE_CLIENT_SECRET = os.getenv("GOOGLE_CLIENT_SECRET", None)
GOOGLE_DISCOVERY_URL = (
    "https://accounts.google.com/.well-known/openid-configuration"
)

google_client = WebApplicationClient(GOOGLE_CLIENT_ID)

def get_google_provider_cfg():
    return requests.get(GOOGLE_DISCOVERY_URL).json()



#FIXME: if ckeditor is enabled, this will not work, I think it has to do with the html tags
#gmail did not give errors..

#send event email to everyone paired
def send_all_event_emails(subject, email_content):
    cleaner = re.compile('<.*?>')
    subject = re.sub(cleaner, '', subject)
    email_content = re.sub(cleaner, '', email_content)
    #print("------send all event emails:")

    #print("--subject: ", subject)
    #print("--email content: ", email_content)

    db_big_answers = Big_Answer.query.filter(Big_Answer.is_paired == True).all()
    db_little_answer = Little_Answer.query.filter(Little_Answer.is_paired == True).all()


    big_to_emails = []
    if db_big_answers:
        for one_big in db_big_answers:
            #one_to_email = [one_big.email]
            big_to_emails.append(one_big.email)

    try:
        send_event_emails(
            from_email=os.getenv("MAIL_USERNAME"),
            to_emails=big_to_emails,
            subject=subject,
            body=email_content
        )
    except Exception as e:
        print("---error sending big email: ", e)

    little_to_emails = []

    if db_little_answer:

        for one_little in db_little_answer:
            #one_to_email = [one_little.email]
            little_to_emails.append(one_little.email)

    try:
        send_event_emails(
            from_email=os.getenv("MAIL_USERNAME"),
            to_emails=little_to_emails,
            subject=subject,
            body=email_content
        )
    except Exception as e:
        print("---error sending little email: ", e)

    #print("--all big emaiL ", big_to_emails)
    #print("--all little emaiL ", little_to_emails)




def read_with_dropdown(book_name, sheet_name="Sheet1", range_str="B2:B100"):
    wb = openpyxl.load_workbook(book_name)
    ws = wb[sheet_name]
    data = [[cell.value for cell in row] for row in ws[range_str]]
    #data = ws["B7"]

    #print("--data is: ", data)

    validations = ws.data_validations.dataValidation
    #print("---the cell: ",validations)

    for validation in validations:
        #print("-------------------------11111---------------------------------")
        #print("--one validation here: ", validation)
        ranges = validation.sqref.ranges
        #ranges = validation.sqref
        #print("--ranges: ", ranges)
        if len(ranges) != 1:
            raise NotImplementedError

        #print("---------2")
        if validation.type == 'list':
            #print("---type is list..")
            #list_cells = ws[validation.formula1]
            list_cells = validation.formula1
            #print("---list cells: ", list_cells)
            #values = [cell.value for cell_row in list_cells for cell in cell_row]
            try:
                values=list_cells.strip().split(",")

                #store a list of idx that have empty anwers
                empty_idx = []
                for i in range(len(values)):
                    #print("--value before:", repr(values[i]))
                    values[i] = values[i].strip().replace('\"', '')
                    #print("--value after:", repr(values[i]))
                    if values[i] == "":
                        empty_idx.append(i)

                #delete the element that is empty string
                for one_idx in empty_idx:
                    values.pop(i)



            except:
                values=[list_cells]

            #print("--a list of options: ", values)
        else:
            raise NotImplementedError

        #print("---------3")
        bounds = ranges[0].bounds
        #print("-bounds: ", bounds)


        #print("---------4")
        try:
            #data[bounds[1]-1][bounds[0]-1] = values

            #this one works,but only record the first one
            #data[bounds[1]-2] = values
            #this should take care all rows that use the same formula
            for i in range(1, len(bounds), 2):
                data[bounds[i]-2] = values[:]

        except IndexError:
            pass


    return data


#returning [[row1], [row2], [row3], [],....]
def paired_littles_to_array(littles_json_list):
    result = []

    questions = []
    db_little_q_set = Little_Question_Set.query.first()
    if not db_little_q_set:
        return []

    little_q_json = db_little_q_set.get_json()
    for key in little_q_json:
        questions.append(little_q_json[key]['question'])

    result.append(questions)


    for one_little_json in littles_json_list:
        one_little_submision = []
        for temp_key in one_little_json:
            if temp_key.isnumeric():
                one_little_submision.append(one_little_json[temp_key]['answer'])

        result.append(one_little_submision[:])

    print("--result is: ", result)

    return result





#for admin.py
#NOT GENERAL: covert question set json in database to list of list to create excel
def q_set_json_to_array(json_data):
    return_list = [["WEIGHT", "Type/Answer", "Question"]]

    for one_key in json_data:
        one_q_info = json_data[one_key]
        weight = one_q_info['weight']
        type_answer = "nan"
        if one_q_info['type'] == 'mc':
            type_answer = ",".join(one_q_info['answers'])

        question = one_q_info['question']

        return_list.append([weight, type_answer, question])

    return return_list


#for main.py
#covert one submission json to what the database expect
def submission_json_to_database_json(q_set_id, raw_json):
    db_question_set = None
    if q_set_id == BIG_QUESTION_SET_ID:
        db_question_set = Big_Question_Set.query.first()
    else:
        db_question_set = Little_Question_Set.query.first()

    questions_json = db_question_set.get_json()

    for q_idx in questions_json:
        try:
            questions_json[q_idx]['answer'] = raw_json[q_idx].strip()
            #questions_json[q_idx]['weight_int']
            weight_raw = questions_json[q_idx]['weight'].lower()
            if weight_raw == 'l':
                questions_json[q_idx]['weight_int'] = LOW
            elif weight_raw == 'm':
                questions_json[q_idx]['weight_int'] = MEDIUM
            elif weight_raw == 'h':
                questions_json[q_idx]['weight_int'] = HIGH
            else:
                questions_json[q_idx]['weight_int'] = NONE


        except:
            #put empty string if no answer is given..
            questions_json[q_idx]['answer'] = ''

    #FIXME: will add more later
    #add some info on top level for easier access
    questions_json['name'] = "N/A"
    questions_json['email'] = "N/A"
    questions_json['age'] = "N/A"
    questions_json['major_dept'] = "N/A"
    questions_json['major'] = "N/A"

    try:
        questions_json['email'] = questions_json['0']['answer']
        questions_json['name'] = questions_json['1']['answer']
        questions_json['age'] = questions_json['2']['answer']
        questions_json['major_dept'] = questions_json['5']['answer']
        questions_json['major'] = questions_json['6']['answer']
    except:
        pass

    print("result json: ", questions_json)

    return questions_json

